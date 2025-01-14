# -*- coding: latin-1 -*-
import datetime
import time
from xmlrpc.client import DateTime
from openpyxl import load_workbook, Workbook
import pymssql
import os
import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from datetime import date, timedelta, datetime
from re import findall
from getpass4 import getpass



class TraxpayScraper:
    def __init__(self, time_range, username, password):
        self.driver = webdriver.Edge()
        self.time_range = time_range
        self.username = username
        self.password = password
       
    def login(self):    
        try:
            self.driver.get('https://financing.traxpay.com/financing/login.xhtml')
            self.driver.maximize_window()
        
            user = self.driver.find_element(By.ID, 'loginForm:inputUsername')
            password = self.driver.find_element(By.ID, 'loginForm:inputPassword')
            login_button = self.driver.find_element(By.ID, 'loginForm:loginButton')

            user.send_keys(self.username)
            password.send_keys(self.password)
        
            login_button.click()
            
            if self.driver.current_url == 'https://financing.traxpay.com/financing/login.xhtml':
                print("Falscher Login!")
                self.quit()
            else:
                print("User eingeloggt.")
                
        except Exception as e:
            print(f"Fehler während des Logins: {e}")
            self.quit()
            
    def go_to_table(self):
        try:
            wait = WebDriverWait(self.driver, 20)
        
            wait.until(EC.presence_of_element_located((By.ID, 'invoiceDtoLazyListModel:j_id_ih')))
            wait.until(EC.element_to_be_clickable((By.ID, 'invoiceDtoLazyListModel:j_id_ih')))
        
            order_by_element = self.driver.find_element(By.ID, 'invoiceDtoLazyListModel:j_id_ih')
        
            self.driver.execute_script("arguments[0].click();", order_by_element)
            time.sleep(2)
            self.driver.execute_script("arguments[0].click();", order_by_element)
        
            print("Tabelle sortiert.")
            time.sleep(2)
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)
            self.driver.find_element(By.XPATH, '//*[@id="invoiceDtoLazyListModel:j_id_jy"]/div[3]').click()
            print('Status auf Alle gesetzt.')
            time.sleep(1)
            self.driver.find_element(By.XPATH, '//*[@id="invoiceDtoLazyListModel:j_id_jy_panel"]/div[1]/div/div[2]/span').click()
            time.sleep(2)
            dropdown_element = self.driver.find_element(By.ID, 'invoiceDtoLazyListModel:j_id__v_1') 
        
            actions = ActionChains(self.driver)
            actions.move_to_element(dropdown_element).perform()
        
            select = Select(dropdown_element)
            select.select_by_visible_text("100")
        
            print("Eintragsanzeige auf 100 gesetzt.")
            time.sleep(3)
            wait.until(EC.presence_of_element_located((By.ID, 'invoiceDtoLazyListModel_data')))
        except Exception as e:
            print(f"Fehler während Tabellennavigation: {e}")
            self.quit()

    def scrape_rglnrs(self):
        try:
            table = self.driver.find_element(By.ID, 'invoiceDtoLazyListModel_data')
            invoice_data = []
            i = 1
            date_range = datetime.today() - timedelta(days=self.time_range)
            paginator = self.driver.find_element(By.XPATH, '//*[@id="invoiceDtoLazyListModel_paginator_bottom"]/span[1]').text
        
            number_of_pages = int(findall(r'\d+', paginator)[-1])
            print(number_of_pages)
    
            for page in range(1, number_of_pages + 1):
                rows = table.find_elements(By.TAG_NAME, 'tr')
                last_entry_date = None
        
                for row in rows:
                    cells = row.find_elements(By.TAG_NAME, 'td')
                    if len(cells) > 0 and len(cells) >= 5:
                        rglnr_span_element = cells[5].find_element(By.TAG_NAME, 'span')
                        rglnr_data = rglnr_span_element.text.strip() if rglnr_span_element else ''
                        date_span_element = cells[7].find_element(By.TAG_NAME, 'span')
                        date_data = date_span_element.text.strip() if date_span_element else ''
                
                        last_entry_date = datetime.strptime(date_data, '%d.%m.%Y')
                
                        if last_entry_date >= date_range:
                            print(f" RGLNR {rglnr_data} erfolgreich scraped --- ({i})")
                            print(f" Datum {date_data} erfolgreich scraped --- ({i})")
                            invoice_data.append((rglnr_data, date_data))
                        elif last_entry_date < date_range:
                            break
                    i += 1
        
                if last_entry_date and last_entry_date < date_range:
                    break
        
                if page < number_of_pages:
                    self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(2)
                    next_page_button = self.driver.find_element(By.XPATH, '//*[@id="invoiceDtoLazyListModel_paginator_bottom"]/a[3]')
                    next_page_button.click()
                    time.sleep(3)
    
            return invoice_data
        except Exception as e:
            print(f"Fehler während des Datenscrapings {e}")
            self.quit()

    def quit(self):
        self.driver.quit()


class MSQueryExecutor:
    def __init__(self, data):
        try:
            self.db = pymssql.connect(server='DC1-DB01', database='wsmb')
        except Exception as e:
            print(f"Datenbank fehler: {e}")
            sys.exit(1)
        self.data = data
        self.rglnr_values = []
        self.grouped_list = []
        
    
    def execute_query(self):
        try:
            rglnr_values = [item[0] for item in self.data]
            cursor = self.db.cursor()
            placeholders = ', '.join(['%s'] * len(rglnr_values))
            query = f"SELECT AXRGNR, RGLNR, ABGDATUMZEIT FROM LOG_RGLNR WHERE RGLNR IN ({placeholders}) ORDER BY ABGDATUMZEIT ASC"
            cursor.execute(query, rglnr_values)
        
            for row in cursor.fetchall():
               formatted_date = row[2].strftime('%d.%m.%Y')
               row2 = row[:2] + (formatted_date,) + row[3:]
               self.grouped_list.append(row2)
           
            return self.grouped_list
    
            cursor.close()
            self.db.close()
        except Exception as e:
            print(f"Fehler während des Ausführens der Datenbankabfrage {e}")
            self.db.close()


class ExcelExport:
        def __init__(self, data, excel_path, save_path):
            self.data = data
            self.excel_path = excel_path
            self.save_path = save_path
            
        def write_to_excel(self):
            try:
                filtered_data = [i for i in self.data]

                print(f"{len(filtered_data)} Eintraege.")
                if not os.path.exists(self.excel_path):
                    print(f"Datei {self.excel_path} existiert nicht. Erstelle neue Datei...")
                    workbook = Workbook()
                    workbook.create_sheet('inTraxpay')
                else:
                    workbook = load_workbook(self.excel_path)
                if 'inTraxpay' not in workbook.sheetnames:
                    workbook.create_sheet('inTraxpay')
            
                sheet = workbook['inTraxpay']
                last_row = sheet.max_row

                for i in filtered_data:
                    last_row += 1
                    for col, value in enumerate(i, start=1):
                        sheet.cell(row=last_row, column=col, value=value)

                workbook.save(self.save_path)
                print('Success')
            except Exception as e:
                print(f"Fehler während des Exportierens der Datei: {e}")
     


if getattr(sys, 'frozen', False) == True:
    application_path = os.path.dirname(sys.executable)
else:
    application_path = os.path.dirname(__file__)


try:
    time_range = int(input("Waehlen Sie den Zeitraum fuer die Datenfilterung: ")) + 1

    username = input("Nutzername: ")
    password = getpass(prompt="Passwort: ")

    scraper = TraxpayScraper(time_range, username, password)
    scraper.login()
    scraper.go_to_table()

    data = scraper.scrape_rglnrs()

    scraper.quit()

    executor = MSQueryExecutor(data)
    grouped_data = executor.execute_query()

    input_filename = input('Dateinamen hier einfuegen: ').strip() or "OP-Liste per 20240524.xlsx"
    excel_path = os.path.join(application_path, input_filename)
    save_path = os.path.join(application_path, "Updated_data.xlsx")

    writer = ExcelExport(grouped_data, excel_path, save_path)
    writer.write_to_excel()
except Exception as e:
    print(f"Ein Fehler ist aufgetreten: {e}")